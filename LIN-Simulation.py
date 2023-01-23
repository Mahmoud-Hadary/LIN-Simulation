import sys
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QTableWidget, QTableWidgetItem, QLabel,QPushButton, QFileDialog
from PyQt5.QtGui import QIcon
import pandas as pd
import xlwt
import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
import csv

# SENDING IS C5D9F1
# RECIEVING IS E6B8B7
coloumn_count = 0
'''
class Anim_Window(QWidget):
    def __init__(self):
        super().__init__()
        self.setGeometry(300, 300, 300, 220)
        self.setWindowTitle('LIN Simulation')
        self.setWindowIcon(QIcon('download.png'))
        self.setFixedSize(610,480)
        self.show()
'''
class LoadTable(QTableWidget,QWidget):
    def switch_window(self):
        self.savefile()
        #self.hide()
        #self.second_window = Anim_Window()
        #self.second_window.show()

    def savefile(self):
        filename = str(QFileDialog.getSaveFileName(self, 'Save File', '', ".csv(*.csv)"))
        filename = filename.split(",")
        filename=filename[0]
        filename=filename[2:-1]
        wbk = xlwt.Workbook()
        self.sheet = wbk.add_sheet("sheet", cell_overwrite_ok=True)
        self.add2()
        if(os.path.exists(filename)):
            os.remove(filename)
        wbk.save(filename)

 

    def anim(self):
        
        global DATa
        global Start_of_COM
        global END_of_COM
        global NUM_of_COL
        try:
            with open('DATA.csv', mode ='r',encoding='cp850')as file:
                # reading the CSV file
                csvFile = csv.reader(file)
                # displaying the contents of the CSV file
                i = 0
                for lines in csvFile:
                    if i == NUM_of_COL:
                        DATa = lines
                    
                    i+=i
        except:
            QtWidgets.QMessageBox.critical(self, "Invalid Input", "There is nothing to animate yet")
        filename = str(QFileDialog.getSaveFileName(self, 'Save File', '', ".xlsx(*.xlsx)"))
        filename = filename.split(",")
        filename=filename[0]
        filename=filename[2:-1]
        wb = Workbook()
        ws = wb.active
        BLACKFill = PatternFill(start_color='FFFFFFFF',
                   end_color='FFFFFFFF',
                   fill_type='solid')
        SENDING = PatternFill(start_color='C5D9F1',
                   end_color='C5D9F1',
                   fill_type='solid')
        RECIEVING = PatternFill(start_color='E6B8B7',
                   end_color='E6B8B7',
                   fill_type='solid')
        ws['A1'].fill = BLACKFill
        ws['A2'].fill = BLACKFill
        ws['A3'].fill = BLACKFill
        ws['A4'].fill = BLACKFill
        ws['A5'].fill = BLACKFill
        ws['A6'].fill = BLACKFill
        ws['A7'].fill = BLACKFill
        ws['A8'].fill = BLACKFill
        ws['A9'].fill = BLACKFill
        ws['A10'].fill = BLACKFill
        ws['B1'].fill = BLACKFill
        ws['C1'].fill = BLACKFill
        ws['C2'].fill = BLACKFill
        ws['C3'].fill = BLACKFill
        ws['C4'].fill = BLACKFill
        ws['C5'].fill = BLACKFill
        ws['C6'].fill = BLACKFill
        ws['C7'].fill = BLACKFill
        ws['C8'].fill = BLACKFill
        ws['C9'].fill = BLACKFill
        ws['C10'].fill = BLACKFill
        ws['C1'].fill = BLACKFill
        ws['B2'] = "SLAVE A"
        ws['B3'] = "Recieve"
        ws['B4'] = "Transmit"
        #C3 for Recieve and C4 for Transmit
        ws['B5'].fill = BLACKFill
        ws['B6'] = "SLAVE B"
        ws['B7'] =  "Recieve"
        ws['B8'] =  "Transmit"
        #C7 for Recieve and C8 for Transmit
        ws['B9'].fill= BLACKFill
        ws['B11'].fill= BLACKFill

        ws['D1'].fill = BLACKFill
        ws['E1'].fill = BLACKFill
        ws['E3'].fill = BLACKFill
        ws['E4'] = "MASTER NODE"
        ws['E5']= "HEADER CREATER"
        ws['E6'].fill =BLACKFill
        ws['E7'].fill =BLACKFill
        ws['E8'].fill =BLACKFill
        ws['E9']= "MASTER RECIEVE"
        ws['E10']= "MASTER TRANSMIT"
        #G9 for Recieve and G10 for Transmit
        ws['E11'].fill= BLACKFill
        ws['F1'].fill = BLACKFill
        ws['F3'].fill = BLACKFill
        ws['F4'].fill= BLACKFill
        ws['F5'].fill= BLACKFill
        ws['F6'].fill =BLACKFill
        ws['F7'].fill =BLACKFill
        ws['F8'].fill =BLACKFill
        ws['F9'].fill= BLACKFill
        ws['F10'].fill= BLACKFill
        ws['F11'].fill= BLACKFill
        print(DATa)

        if(os.path.exists(filename)):
            os.remove(filename)
        wb.save(filename) 
        NUM_of_COL+=NUM_of_COL
        #for j in range(self.rowCount()):
            #for i in range(self.columnCount()):
                
    def add2(self):
        row = 0
        col = 0         
        for i in range(self.columnCount()):
            for x in range(self.rowCount()):   
                try:     
                    teext = str(self.item(row, col).text())
                    if col > 0:
                        if teext != "Recieved" and teext != "Dont Care" and teext != "Transfer":
                            QtWidgets.QMessageBox.critical(self, "Invalid Input", "The Input for the Nodes should be either Recieved, Dont Care, or Transfer")
                    self.sheet.write(row, col, teext)
                    row += 1
                
                except AttributeError:
                    row += 1
            row = 0
            col +=1
        #content = self.combo_box.currentText()
    def __init__(self, parent=None):
        
        super(LoadTable, self).__init__(1, 4, parent)
        self.setFont(QtGui.QFont("Helvetica", 10, QtGui.QFont.Normal, italic=False))   
        headertitle = ("Frame ID","Master","Slave A","Slave B")
        self.setHorizontalHeaderLabels(headertitle)
        self.verticalHeader().hide()
        self.horizontalHeader().setHighlightSections(False)
        self.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Fixed)

        self.setSelectionMode(QtWidgets.QAbstractItemView.NoSelection)
        self.setColumnWidth(0, 130)

        self.cellChanged.connect(self._cellclicked)
        


    @QtCore.pyqtSlot(int, int)
    def _cellclicked(self, r, c):
        it = self.item(r, c)
        it.setTextAlignment(QtCore.Qt.AlignCenter)        

    @QtCore.pyqtSlot()
    def _addrow(self):
        rowcount = self.rowCount()
        self.insertRow(rowcount)
 

    @QtCore.pyqtSlot()
    def _removerow(self):
        if self.rowCount() > 0:
            self.removeRow(self.rowCount()-1)


class ThirdTabLoads(QWidget):

    def __init__(self, parent=None):
        super(ThirdTabLoads, self).__init__(parent)    
        self.setFixedSize(560,450)
        self.setWindowTitle('LIN Simulation')
        self.setWindowIcon(QIcon('download.png'))
        table = LoadTable()
        #layout = QVBoxLayout()
        #self.setLayout(layout)

        add_button = QtWidgets.QPushButton("Add")
        add_button.clicked.connect(table._addrow)
        
        Save_button = QtWidgets.QPushButton("Save")
        #self.Save_button = QPushButton('&Export',clicked=self.switch_window)
        #layout.addWidget(self.Save_button)
        Save_button.clicked.connect(table.savefile)  #saves the data and transfers it into an excel sheet then opens a new window

        delete_button = QtWidgets.QPushButton("Delete")
        delete_button.clicked.connect(table._removerow)

        CON_button = QtWidgets.QPushButton("Animate")
        CON_button.clicked.connect(table.anim)
        button_layout = QtWidgets.QVBoxLayout()
        button_layout.addWidget(add_button, alignment=QtCore.Qt.AlignBottom)
        button_layout.addWidget(delete_button, alignment=QtCore.Qt.AlignTop)
        button_layout.addWidget(Save_button, alignment=QtCore.Qt.AlignBottom)
        try:
            button_layout.addWidget(CON_button, alignment=QtCore.Qt.AlignBottom)
        except:
            QtWidgets.QMessageBox.critical(self, "Invalid Input", "There is nothing to animate yet")
        tablehbox = QtWidgets.QHBoxLayout()
        tablehbox.setContentsMargins(10, 10, 10, 10)
        tablehbox.addWidget(table)

        grid = QtWidgets.QGridLayout(self)
        grid.addLayout(button_layout, 0, 1)
        grid.addLayout(tablehbox, 0, 0)        


if __name__ == '__main__':
    Start_of_COM = False
    END_of_COM = True
    DATa=""
    NUM_of_COL=0
    app = QtWidgets.QApplication(sys.argv)
    
    w = ThirdTabLoads()
    
    w.show()
    sys.exit(app.exec_())